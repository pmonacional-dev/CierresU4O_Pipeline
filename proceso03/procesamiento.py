#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
sys.dont_write_bytecode = True
r"""
Pipeline de construcción de tablas para modeloDatos.xlsx

Crea/actualiza estas hojas/tablas:
1) Ventas24-25         (TablaVentas24-25)
2) Ventas25-26         (TablaVentas25-26)
3) VentasRegion        (TablaVentasRegion)
4) VentasGeneral       (TablaVentasGeneral)         -> TODAS las zonas (incluye INTERNACIONAL)
5) VentasNacional      (TablaVentasNacional)        -> SOLO nacionales (CENTRO & OCCIDENTE, MÉXICO, MONTERREY, NOROESTE)
6) VentasInternacional (TablaVentasInternacional)   -> SOLO INTERNACIONAL (USA ya sumado)
7) EmbudoGeneral       (TablaEmbudoGeneral)
8) EmbudoRegion        (TablaEmbudoRegion)

Metas:
- Tomamos la meta acumulada (general) de la hoja "Metas".
- Para NACIONAL e INTERNACIONAL, escalamos semana a semana usando estos totales:
    General:       841,320,000
    Nacional:      665,760,000
    Internacional: 175,560,000
"""

import os
import time
import shutil
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os, time, shutil
from pathlib import Path

# Variables que puede inyectar generar_modelodatos(...)
BASE_DIR: str | None = None
SRC_PATH: str | None = None
DST_PATH: str | None = None
SRC_SHEET: str | None = None


def copy_with_retries(src: str, dst: str, attempts: int = 8, delay: float = 1.5):
    """Copia src → dst con reintentos (útil si OneDrive bloquea un instante el archivo)."""
    last_err = None
    for _ in range(attempts):
        try:
            Path(dst).parent.mkdir(parents=True, exist_ok=True)
            tmp_dst = dst + ".tmp"
            shutil.copyfile(src, tmp_dst)
            if os.path.exists(dst):
                os.remove(dst)
            os.replace(tmp_dst, dst)
            return True
        except Exception as e:
            last_err = e
            time.sleep(delay)
    raise last_err


SRC_FILENAME = "Tableau_WinRoom_ImagenFunnel_PBI.xlsx"
DST_FILENAME = "modeloDatos.xlsx"
SRC_SHEET = "Result 1"


ZONAS_ORDEN = ["CENTRO & OCCIDENTE", "MÉXICO", "MONTERREY", "NOROESTE", "INTERNACIONAL"]
ZONAS_NACIONALES = ["CENTRO & OCCIDENTE", "MÉXICO", "MONTERREY", "NOROESTE"]  # para Nacional
ZONA_ORDER_MAP = {z: i for i, z in enumerate(ZONAS_ORDEN)}

# Totales de meta del periodo (proporcionados)
TOTAL_META_GENERAL = 841_320_000.00
TOTAL_META_NACIONAL = 665_760_000.00
TOTAL_META_INTERNACIONAL = 175_560_000.00

MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
            7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

# ---------------------- HELPERS ----------------------
def remove_accents(s: str) -> str:
    """Quita acentos y eñes; útil para comparaciones insensibles a acentos."""
    if s is None:
        return ""
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),
                 ("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),
                 ("ñ","n"),("Ñ","N")]:
        s = s.replace(a, b)
    return s

def norm_region_label(label: str) -> str:
    """Normaliza etiquetas de zona; mapea USA→INTERNACIONAL y MEXICO→MÉXICO."""
    if label is None:
        return ""
    t = str(label).strip()
    if t.upper() == "USA":
        return "INTERNACIONAL"
    if remove_accents(t).upper() == "MEXICO":
        return "MÉXICO"
    return t

def normalize_etapa(e):
    """Normaliza etapa a minúsculas sin acentos (para comparaciones/llaves)."""
    if e is None:
        return ""
    return remove_accents(str(e)).strip().lower()

def etapa_label(e_norm: str) -> str:
    """Devuelve la etiqueta 'bonita' de la etapa ya normalizada."""
    m = {"oportunidad": "Oportunidad", "propuesta": "Propuesta",
         "negociacion": "Negociación", "cerrada ganada": "Cerrada Ganada"}
    return m.get(e_norm, str(e_norm).title())

def to_number(x):
    """Convierte strings tipo $1,234 o (1,234) a float."""
    if pd.isna(x):
        return float("nan")
    s = str(x).strip()
    if s == "":
        return float("nan")
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    try:
        v = float(s)
        return -v if neg else v
    except:
        return float("nan")

def crear_o_reemplazar_tabla(wb, sheet_name: str, table_name: str):
    """Crea/actualiza una Excel Table que cubre todos los datos de la hoja indicada."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 1 or max_col < 1:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    # Eliminar si ya existía
    to_remove = None
    for t in ws._tables:
        if t.displayName == table_name:
            to_remove = t
            break
    if to_remove:
        ws._tables.remove(to_remove)
    # Crear de nuevo con estilo
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def copy_to_temp(path: str) -> str:
    """Hace una copia temporal del archivo para operar sin bloqueos."""
    work_dir = os.path.join(tempfile.gettempdir(), "modelo_datos_work_pipeline")
    os.makedirs(work_dir, exist_ok=True)
    work_path = os.path.join(work_dir, "modeloDatos_work.xlsx")
    shutil.copy2(path, work_path)
    return work_path

def copy_back_with_retries(src: str, dst: str, retries: int = 20, wait: float = 0.5) -> bool:
    """Copia/reemplaza con reintentos para esquivar bloqueos de OneDrive/Excel."""
    last = None
    for _ in range(retries):
        try:
            try:
                os.replace(src, dst)
            except PermissionError:
                shutil.copy2(src, dst)
            return True
        except PermissionError as e:
            last = e
            time.sleep(wait)
    if last:
        raise last
    return False

def get_semana_column(df):
    """Localiza la columna de semana en 'Metas' (por nombre que contenga 'semana')."""
    for c in df.columns:
        n = remove_accents(str(c)).strip().lower()
        if "semana" in n:
            return c
    return df.columns[0]

def get_acumulado_column(df):
    """Localiza la columna 'Acumulado' (para metas generales)."""
    for c in df.columns:
        n = remove_accents(str(c)).strip().lower()
        if n == "acumulado":
            return c
    # fallback
    return df.columns[3] if len(df.columns) >= 4 else df.columns[-1]

def _mode_month(s):
    """Moda del mes numérico (rompe empates con el menor)."""
    vc = s.dropna().astype(int).value_counts()
    if vc.empty:
        return pd.NA
    maxc = vc.max()
    candidates = vc[vc == maxc].index.tolist()
    return int(min(candidates))

# ---------------------- CONSTRUCCIÓN: Ventas por periodo ----------------------
def preparar_ventas_por_periodo(df_src, periodo_str):
    """
    Filtra Periodo=periodo_str y Etapa='Cerrada Ganada', normaliza zona (USA→INTERNACIONAL),
    agrupa por Zona+NoSemana sumando Importe y calcula DiferenciaSemanal intra-zona.
    """
    d = df_src[df_src["Periodo"].astype(str).str.strip() == periodo_str].copy()
    d["_etapa_norm"] = d["Etapa"].apply(normalize_etapa)
    d = d[d["_etapa_norm"] == "cerrada ganada"].copy()
    d["ZonaAgrupada"] = d["ZonaAgrupada"].apply(norm_region_label)
    d["NoSemana"] = pd.to_numeric(d["NoSemana"], errors="coerce")
    d["Importe"] = pd.to_numeric(d["Importe"], errors="coerce").fillna(0.0)

    g = d.groupby(["ZonaAgrupada", "NoSemana"], as_index=False, dropna=False)["Importe"].sum()

    # Orden consistente por zona
    g["__zona_ord"] = g["ZonaAgrupada"].map(ZONA_ORDER_MAP).fillna(9999)
    g = g.sort_values(["__zona_ord", "NoSemana"]).drop(columns="__zona_ord").reset_index(drop=True)

    # DiferenciaSemanal (primera semana = mismo valor)
    g["DiferenciaSemanal"] = g.groupby("ZonaAgrupada")["Importe"].diff()
    first_idx = g.groupby("ZonaAgrupada").head(1).index
    g.loc[first_idx, "DiferenciaSemanal"] = g.loc[first_idx, "Importe"]

    return g[["NoSemana", "ZonaAgrupada", "Importe", "DiferenciaSemanal"]]

# ---------------------- METAS: por región y generales ----------------------
def metas_por_region_desde_metas(df_metas):
    """
    Parsea hoja 'Metas' cuando el layout va en pares [REGIÓN][VALORES acumulados].
    Devuelve: NoSemana, ZonaAgrupada, metaSemanal (acumulado), metaDesgloce (semana 1).
    - Suma USA dentro de INTERNACIONAL.
    - Omite GOBIERNO si aparece.
    """
    dfm = df_metas.copy()
    dfm.columns = [str(c).strip() for c in dfm.columns]
    semana_col = get_semana_column(dfm)
    dfm = dfm.rename(columns={semana_col: "NoSemana"})
    dfm["NoSemana"] = pd.to_numeric(dfm["NoSemana"], errors="coerce")

    cols = list(dfm.columns)
    region_series = {}
    for idx, col in enumerate(cols):
        if col == "NoSemana":
            continue
        region_name_norm = norm_region_label(col)
        # ¿Esta columna es encabezado de región?
        if (region_name_norm in ZONAS_ORDEN) or (region_name_norm in {"USA", "GOBIERNO"}):
            next_idx = idx + 1
            if next_idx < len(cols):
                valores_col = cols[next_idx]
                maybe_region_next = norm_region_label(valores_col)
                # Si el siguiente también es región, no es un par válido
                if (maybe_region_next in ZONAS_ORDEN) or (maybe_region_next in {"USA", "GOBIERNO"}):
                    continue
                serie = dfm[valores_col].apply(to_number)
                region_series[region_name_norm] = serie

    # INTERNACIONAL = INTERNACIONAL + USA si existe
    internacional_acum = None
    if "INTERNACIONAL" in region_series:
        internacional_acum = region_series["INTERNACIONAL"].copy()
    if "USA" in region_series:
        if internacional_acum is not None:
            internacional_acum = pd.to_numeric(internacional_acum, errors="coerce").fillna(0) + \
                                 pd.to_numeric(region_series["USA"], errors="coerce").fillna(0)
        else:
            internacional_acum = pd.to_numeric(region_series["USA"], errors="coerce")

    region_series.pop("GOBIERNO", None)

    metas_list = []
    for z in ZONAS_ORDEN:
        if z == "INTERNACIONAL":
            if internacional_acum is None:
                continue
            acum = internacional_acum
        else:
            if z not in region_series:
                continue
            acum = region_series[z]
        tmp = pd.DataFrame({
            "NoSemana": dfm["NoSemana"],
            "ZonaAgrupada": z,
            "metaSemanal": pd.to_numeric(acum, errors="coerce")
        })
        first_val = tmp.sort_values("NoSemana").iloc[0]["metaSemanal"] if not tmp.empty else float("nan")
        tmp["metaDesgloce"] = first_val
        metas_list.append(tmp)

    if not metas_list:
        return pd.DataFrame(columns=["NoSemana", "ZonaAgrupada", "metaSemanal", "metaDesgloce"])
    return pd.concat(metas_list, ignore_index=True)

def metas_generales_desde_acumulado(df_metas):
    """Toma 'Acumulado' como metaSemanal general; metaDesgloce = semana 1."""
    df = df_metas.copy()
    df.columns = [str(c).strip() for c in df.columns]
    semana_col = get_semana_column(df)
    acum_col = get_acumulado_column(df)
    out = df[[semana_col, acum_col]].rename(columns={semana_col: "NoSemana", acum_col: "metaSemanal"})
    out["NoSemana"] = pd.to_numeric(out["NoSemana"], errors="coerce")
    out["metaSemanal"] = out["metaSemanal"].apply(to_number)
    meta_w1 = out.sort_values("NoSemana").iloc[0]["metaSemanal"] if not out.empty else float("nan")
    out["metaDesgloce"] = meta_w1
    return out

# ---------------------- EMBUDOS ----------------------
ETAPAS_OBJ = ["oportunidad", "propuesta", "negociacion", "cerrada ganada"]
FACTOR_ETAPA = {"oportunidad": 0.30, "propuesta": 0.50, "negociacion": 0.80, "cerrada ganada": 1.00}
ETAPA_ORDER = {"Oportunidad": 0, "Propuesta": 1, "Negociación": 2, "Cerrada Ganada": 3}

def preparar_embudo_general(df_src):
    """Embudo por semana y etapa (sin separar regiones)."""
    df = df_src.copy()
    df["NoSemana"] = pd.to_numeric(df["NoSemana"], errors="coerce")
    df["Importe"] = pd.to_numeric(df["Importe"], errors="coerce").fillna(0.0)
    df["TotalProyectos"] = pd.to_numeric(df.get("TotalProyectos", 0), errors="coerce").fillna(0.0)
    df["FechaCierre_dt"] = pd.to_datetime(df.get("FechaCierre", pd.NaT), errors="coerce")
    df["MesNum"] = df["FechaCierre_dt"].dt.month
    df["_etapa_norm"] = df["Etapa"].apply(normalize_etapa)
    df["ZonaAgrupada"] = df["ZonaAgrupada"].apply(norm_region_label)

    # Periodo actual
    cur = df[(df["Periodo"].astype(str).str.strip() == "25-26") & (df["_etapa_norm"].isin(ETAPAS_OBJ))].copy()
    cur["_etapa_label"] = cur["_etapa_norm"].apply(etapa_label)
    agg_cur = cur.groupby(["NoSemana", "_etapa_label"], as_index=False).agg(
        TotalProyectos=("TotalProyectos", "sum"),
        Importe=("Importe", "sum"),
        MesNum=("MesNum", _mode_month)
    )
    # Valor esperado
    def factor(lbl):
        return FACTOR_ETAPA[remove_accents(lbl).lower()]
    agg_cur["ValorEsperado"] = agg_cur.apply(lambda r: r["Importe"] * factor(r["_etapa_label"]), axis=1)
    agg_cur["mes"] = agg_cur["MesNum"].map(MESES_ES)

    # Periodo anterior
    prev = df[(df["Periodo"].astype(str).str.strip() == "24-25") & (df["_etapa_norm"].isin(ETAPAS_OBJ))].copy()
    prev["_etapa_label"] = prev["_etapa_norm"].apply(etapa_label)
    agg_prev = prev.groupby(["NoSemana", "_etapa_label"], as_index=False).agg(PeriodoAnterior=("Importe", "sum"))

    emb = agg_cur.merge(agg_prev, on=["NoSemana", "_etapa_label"], how="left")
    emb["__ord"] = emb["_etapa_label"].map(ETAPA_ORDER).fillna(99)
    emb = emb.sort_values(["__ord", "NoSemana"]).drop(columns="__ord").reset_index(drop=True)

    return emb.rename(columns={"_etapa_label": "Etapa"})[
        ["NoSemana", "mes", "Etapa", "TotalProyectos", "Importe", "ValorEsperado", "PeriodoAnterior"]
    ]

def preparar_embudo_region(df_src):
    """Embudo por semana, región y etapa (USA unida a INTERNACIONAL)."""
    df = df_src.copy()
    df["NoSemana"] = pd.to_numeric(df["NoSemana"], errors="coerce")
    df["Importe"] = pd.to_numeric(df["Importe"], errors="coerce").fillna(0.0)
    df["TotalProyectos"] = pd.to_numeric(df.get("TotalProyectos", 0), errors="coerce").fillna(0.0)
    df["FechaCierre_dt"] = pd.to_datetime(df.get("FechaCierre", pd.NaT), errors="coerce")
    df["MesNum"] = df["FechaCierre_dt"].dt.month
    df["_etapa_norm"] = df["Etapa"].apply(normalize_etapa)
    df["ZonaAgrupada"] = df["ZonaAgrupada"].apply(norm_region_label)
    df = df[df["ZonaAgrupada"].isin(ZONAS_ORDEN)]  # limitar a las zonas esperadas

    # Periodo actual
    cur = df[(df["Periodo"].astype(str).str.strip() == "25-26") & (df["_etapa_norm"].isin(ETAPAS_OBJ))].copy()
    cur["_etapa_label"] = cur["_etapa_norm"].apply(etapa_label)
    agg_cur = cur.groupby(["NoSemana", "ZonaAgrupada", "_etapa_label"], as_index=False).agg(
        TotalProyectos=("TotalProyectos", "sum"),
        Importe=("Importe", "sum"),
        MesNum=("MesNum", _mode_month)
    )
    def factor(lbl):
        return FACTOR_ETAPA[remove_accents(lbl).lower()]
    agg_cur["ValorEsperado"] = agg_cur.apply(lambda r: r["Importe"] * factor(r["_etapa_label"]), axis=1)
    agg_cur["mes"] = agg_cur["MesNum"].map(MESES_ES)

    # Periodo anterior
    prev = df[(df["Periodo"].astype(str).str.strip() == "24-25") & (df["_etapa_norm"].isin(ETAPAS_OBJ))].copy()
    prev["_etapa_label"] = prev["_etapa_norm"].apply(etapa_label)
    agg_prev = prev.groupby(["NoSemana", "ZonaAgrupada", "_etapa_label"], as_index=False).agg(PeriodoAnterior=("Importe", "sum"))

    emb = agg_cur.merge(agg_prev, on=["NoSemana", "ZonaAgrupada", "_etapa_label"], how="left")
    emb["__zona_ord"] = emb["ZonaAgrupada"].map(ZONA_ORDER_MAP).fillna(9999)
    emb["__et_ord"] = emb["_etapa_label"].map(ETAPA_ORDER).fillna(99)
    emb = emb.sort_values(["__zona_ord", "__et_ord", "NoSemana"]).drop(columns=["__zona_ord", "__et_ord"]).reset_index(drop=True)

    return emb.rename(columns={"_etapa_label": "Etapa"})[
        ["NoSemana", "mes", "ZonaAgrupada", "Etapa", "TotalProyectos", "Importe", "ValorEsperado", "PeriodoAnterior"]
    ]

def preparar_embudo_general_nacional(df_src):
    """Embudo por semana y etapa EXCLUYENDO la zona INTERNACIONAL (USA ya unida ahí).
    Calcula exactamente igual que preparar_embudo_general.
    """
    df = df_src.copy()

    # Tipos y normalizaciones (mismo patrón que preparar_embudo_general)
    df["NoSemana"] = pd.to_numeric(df["NoSemana"], errors="coerce")
    df["Importe"] = pd.to_numeric(df["Importe"], errors="coerce").fillna(0.0)
    df["TotalProyectos"] = pd.to_numeric(df.get("TotalProyectos", 0), errors="coerce").fillna(0.0)
    df["FechaCierre_dt"] = pd.to_datetime(df.get("FechaCierre", pd.NaT), errors="coerce")
    df["MesNum"] = df["FechaCierre_dt"].dt.month
    df["_etapa_norm"] = df["Etapa"].apply(normalize_etapa)
    df["ZonaAgrupada"] = df["ZonaAgrupada"].apply(norm_region_label)

    # *** Solo nacionales: excluir INTERNACIONAL (donde ya viene unido con USA) ***
    df = df[df["ZonaAgrupada"].str.upper() != "INTERNACIONAL"]

    # ---- Periodo actual (25-26) ----
    cur = df[
        (df["Periodo"].astype(str).str.strip() == "25-26")
        & (df["_etapa_norm"].isin(ETAPAS_OBJ))
    ].copy()
    cur["_etapa_label"] = cur["_etapa_norm"].apply(etapa_label)
    agg_cur = cur.groupby(["NoSemana", "_etapa_label"], as_index=False).agg(
        TotalProyectos=("TotalProyectos", "sum"),
        Importe=("Importe", "sum"),
        MesNum=("MesNum", _mode_month)
    )

    # Valor esperado (mismo factor por etapa)
    def _factor(lbl):
        return FACTOR_ETAPA[remove_accents(lbl).lower()]
    agg_cur["ValorEsperado"] = agg_cur.apply(
        lambda r: r["Importe"] * _factor(r["_etapa_label"]), axis=1
    )
    agg_cur["mes"] = agg_cur["MesNum"].map(MESES_ES)

    # ---- Periodo anterior (24-25) ----
    prev = df[
        (df["Periodo"].astype(str).str.strip() == "24-25")
        & (df["_etapa_norm"].isin(ETAPAS_OBJ))
    ].copy()
    prev["_etapa_label"] = prev["_etapa_norm"].apply(etapa_label)
    agg_prev = prev.groupby(["NoSemana", "_etapa_label"], as_index=False).agg(
        PeriodoAnterior=("Importe", "sum")
    )

    # ---- Unión y orden ----
    emb = agg_cur.merge(agg_prev, on=["NoSemana", "_etapa_label"], how="left")
    emb["__ord"] = emb["_etapa_label"].map(ETAPA_ORDER).fillna(99)
    emb = emb.sort_values(["__ord", "NoSemana"]).drop(columns="__ord").reset_index(drop=True)

    # ---- Salida con las mismas columnas que EmbudoGeneral ----
    return emb.rename(columns={"_etapa_label": "Etapa"})[
        ["NoSemana", "mes", "Etapa", "TotalProyectos", "Importe", "ValorEsperado", "PeriodoAnterior"]
    ]

# ---------------------- MAIN ----------------------
def main():
    # Al inicio de main()
    import os

    # Toma las rutas inyectadas por generar_modelodatos(...) o aplica defaults
    base_dir = globals().get("BASE_DIR", os.getcwd())
    src  = globals().get("SRC_PATH")  or os.path.join(base_dir, "Tableau_WinRoom_ImagenFunnel_PBI.xlsx")
    sheet = globals().get("SRC_SHEET", "Result 1")
    dst  = globals().get("DST_PATH")  or os.path.join(base_dir, "modeloDatos.xlsx")

    print(f"[procesamiento] SRC_PATH = {src}")
    print(f"[procesamiento] SRC_SHEET = {sheet}")
    print(f"[procesamiento] DST_PATH = {dst}")

    # Validación de archivos
    # Validación de archivos
    if not os.path.exists(src):
        raise FileNotFoundError(f"No se encontró el archivo fuente: {src}")

    # Si el destino no existe, créalo vacío para poder abrirlo en modo "a"
    if not os.path.exists(dst):
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        with pd.ExcelWriter(dst, engine="openpyxl", mode="w") as writer:
            # opcional: deja una hoja mínima para que exista el archivo
            pd.DataFrame({"_init": [1]}).to_excel(writer, index=False, sheet_name="init")

    # Leer fuente
    df_src = pd.read_excel(src, sheet_name=sheet, engine="openpyxl")
    df_src.columns = [c.strip() for c in df_src.columns]
    requeridas = ["NoSemana", "Periodo", "ZonaAgrupada", "Etapa", "Importe"]
    faltantes = [c for c in requeridas if c not in df_src.columns]
    if faltantes:
        raise ValueError(f"Faltan columnas requeridas en '{SRC_SHEET}': {faltantes}")

    # Construir ventas por periodo
    ventas_2425 = preparar_ventas_por_periodo(df_src, "24-25")
    ventas_2526 = preparar_ventas_por_periodo(df_src, "25-26")

    # Copia temporal del destino
    work_path = copy_to_temp(dst)


    # Escribir Ventas24-25 y Ventas25-26
    with pd.ExcelWriter(work_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        ventas_2425.to_excel(writer, sheet_name="Ventas24-25", index=False)
        ventas_2526.to_excel(writer, sheet_name="Ventas25-26", index=False)

    wb = load_workbook(work_path)
    # Si existe la hoja 'init' (que usamos para crear el archivo vacío), elimínala
    if "init" in wb.sheetnames and len(wb.sheetnames) > 1:
        ws = wb["init"]
        wb.remove(ws)

    crear_o_reemplazar_tabla(wb, "Ventas24-25", "TablaVentas24-25")
    crear_o_reemplazar_tabla(wb, "Ventas25-26", "TablaVentas25-26")
    wb.save(work_path)

    # Leer Metas
    try:
        metas_raw = pd.read_excel(work_path, sheet_name="Metas", engine="openpyxl")

    except Exception:
        metas_raw = pd.read_excel(work_path, sheet_name="metas", engine="openpyxl")

    # ---------------- VentasRegion (incluye INTERNACIONAL con USA) ----------------
    base = pd.concat([ventas_2425[["NoSemana", "ZonaAgrupada"]],
                      ventas_2526[["NoSemana", "ZonaAgrupada"]]], ignore_index=True).drop_duplicates()
    base = base[base["ZonaAgrupada"].isin(ZONAS_ORDEN)]
    v2425 = ventas_2425.rename(columns={"Importe": "venta24-25", "DiferenciaSemanal": "venta24-25desgloce"}) \
                       [["NoSemana", "ZonaAgrupada", "venta24-25", "venta24-25desgloce"]]
    v2526 = ventas_2526.rename(columns={"Importe": "venta25-26", "DiferenciaSemanal": "venta25-26desgloce"}) \
                       [["NoSemana", "ZonaAgrupada", "venta25-26", "venta25-26desgloce"]]
    region = base.merge(v2425, on=["NoSemana", "ZonaAgrupada"], how="left") \
                 .merge(v2526, on=["NoSemana", "ZonaAgrupada"], how="left")

    metas_region = metas_por_region_desde_metas(metas_raw)
    region = region.merge(metas_region, on=["NoSemana", "ZonaAgrupada"], how="left")

    region["PorcentajeMeta"] = (region["venta25-26"] / region["metaSemanal"] * 100).where(region["metaSemanal"].ne(0))
    region["PorcentajeMetaDesgloce"] = (region["venta25-26desgloce"] / region["metaDesgloce"] * 100).where(region["metaDesgloce"].ne(0))

    region = region[[
        "NoSemana", "ZonaAgrupada", "venta24-25", "venta25-26", "metaSemanal",
        "venta24-25desgloce", "venta25-26desgloce", "metaDesgloce",
        "PorcentajeMeta", "PorcentajeMetaDesgloce"
    ]].sort_values(["ZonaAgrupada", "NoSemana"])

    # ---------------- Metas generales (para General/Nacional/Internacional) ----------------
    metas_general = metas_generales_desde_acumulado(metas_raw).sort_values("NoSemana")
    # Tomar el total de la última semana como meta total general (si no, usar constante proporcionada)
    if not metas_general.empty and metas_general["metaSemanal"].notna().any():
        meta_total_general = metas_general["metaSemanal"].dropna().iloc[-1]
    else:
        meta_total_general = TOTAL_META_GENERAL
    if not meta_total_general or meta_total_general == 0:
        meta_total_general = TOTAL_META_GENERAL

    # ---------------- VentasGeneral (todas las zonas) ----------------
    g2425_all = ventas_2425.groupby("NoSemana", as_index=False)["Importe"].sum().rename(columns={"Importe": "venta24-25"})
    g2526_all = ventas_2526.groupby("NoSemana", as_index=False)["Importe"].sum().rename(columns={"Importe": "venta25-26"})

    base_weeks_all = pd.concat([g2425_all[["NoSemana"]], g2526_all[["NoSemana"]]], ignore_index=True).drop_duplicates().sort_values("NoSemana")

    g2425_all = g2425_all.sort_values("NoSemana"); g2526_all = g2526_all.sort_values("NoSemana")
    g2425_all["24-25desgloce"] = g2425_all["venta24-25"].diff()
    g2526_all["25-26desgloce"] = g2526_all["venta25-26"].diff()
    if not g2425_all.empty: g2425_all.loc[g2425_all.index.min(), "24-25desgloce"] = g2425_all.iloc[0]["venta24-25"]
    if not g2526_all.empty: g2526_all.loc[g2526_all.index.min(), "25-26desgloce"] = g2526_all.iloc[0]["venta25-26"]

    vg_all = base_weeks_all.merge(g2425_all, on="NoSemana", how="left") \
                           .merge(g2526_all, on="NoSemana", how="left") \
                           .merge(metas_general, on="NoSemana", how="left")

    vg_all["porcentajeMeta"] = (vg_all["venta25-26"] / vg_all["metaSemanal"] * 100).where(vg_all["metaSemanal"].ne(0))
    vg_all["porcentajeMetaDesgloce"] = (vg_all["25-26desgloce"] / vg_all["metaDesgloce"] * 100).where(vg_all["metaDesgloce"].ne(0))
    vg_all = vg_all[[
        "NoSemana", "venta24-25", "venta25-26", "metaSemanal",
        "24-25desgloce", "25-26desgloce", "metaDesgloce",
        "porcentajeMeta", "porcentajeMetaDesgloce"
    ]].sort_values("NoSemana")

    # ---------------- VentasNacional (solo nacionales) ----------------
    v2425_nat = ventas_2425[ventas_2425["ZonaAgrupada"].isin(ZONAS_NACIONALES)]
    v2526_nat = ventas_2526[ventas_2526["ZonaAgrupada"].isin(ZONAS_NACIONALES)]
    g2425_nat = v2425_nat.groupby("NoSemana", as_index=False)["Importe"].sum().rename(columns={"Importe": "venta24-25"})
    g2526_nat = v2526_nat.groupby("NoSemana", as_index=False)["Importe"].sum().rename(columns={"Importe": "venta25-26"})

    base_weeks_nat = pd.concat([g2425_nat[["NoSemana"]], g2526_nat[["NoSemana"]]], ignore_index=True).drop_duplicates().sort_values("NoSemana")

    g2425_nat = g2425_nat.sort_values("NoSemana"); g2526_nat = g2526_nat.sort_values("NoSemana")
    g2425_nat["24-25desgloce"] = g2425_nat["venta24-25"].diff()
    g2526_nat["25-26desgloce"] = g2526_nat["venta25-26"].diff()
    if not g2425_nat.empty: g2425_nat.loc[g2425_nat.index.min(), "24-25desgloce"] = g2425_nat.iloc[0]["venta24-25"]
    if not g2526_nat.empty: g2526_nat.loc[g2526_nat.index.min(), "25-26desgloce"] = g2526_nat.iloc[0]["venta25-26"]

    # Escalar metas generales a metas nacionales
    factor_nat = TOTAL_META_NACIONAL / meta_total_general
    metas_nat = metas_general.copy()
    metas_nat["metaSemanal"] = metas_nat["metaSemanal"] * factor_nat
    metas_nat["metaDesgloce"] = metas_nat["metaDesgloce"] * factor_nat

    vg_nat = base_weeks_nat.merge(g2425_nat, on="NoSemana", how="left") \
                           .merge(g2526_nat, on="NoSemana", how="left") \
                           .merge(metas_nat, on="NoSemana", how="left")

    vg_nat["porcentajeMeta"] = (vg_nat["venta25-26"] / vg_nat["metaSemanal"] * 100).where(vg_nat["metaSemanal"].ne(0))
    vg_nat["porcentajeMetaDesgloce"] = (vg_nat["25-26desgloce"] / vg_nat["metaDesgloce"] * 100).where(vg_nat["metaDesgloce"].ne(0))
    vg_nat = vg_nat[[
        "NoSemana", "venta24-25", "venta25-26", "metaSemanal",
        "24-25desgloce", "25-26desgloce", "metaDesgloce",
        "porcentajeMeta", "porcentajeMetaDesgloce"
    ]].sort_values("NoSemana")

    # ---------------- VentasInternacional (solo INTERNACIONAL) ----------------
    v2425_int = ventas_2425[ventas_2425["ZonaAgrupada"] == "INTERNACIONAL"]
    v2526_int = ventas_2526[ventas_2526["ZonaAgrupada"] == "INTERNACIONAL"]
    g2425_int = v2425_int.groupby("NoSemana", as_index=False)["Importe"].sum().rename(columns={"Importe": "venta24-25"})
    g2526_int = v2526_int.groupby("NoSemana", as_index=False)["Importe"].sum().rename(columns={"Importe": "venta25-26"})

    base_weeks_int = pd.concat([g2425_int[["NoSemana"]], g2526_int[["NoSemana"]]], ignore_index=True).drop_duplicates().sort_values("NoSemana")

    g2425_int = g2425_int.sort_values("NoSemana"); g2526_int = g2526_int.sort_values("NoSemana")
    g2425_int["24-25desgloce"] = g2425_int["venta24-25"].diff()
    g2526_int["25-26desgloce"] = g2526_int["venta25-26"].diff()
    if not g2425_int.empty: g2425_int.loc[g2425_int.index.min(), "24-25desgloce"] = g2425_int.iloc[0]["venta24-25"]
    if not g2526_int.empty: g2526_int.loc[g2526_int.index.min(), "25-26desgloce"] = g2526_int.iloc[0]["venta25-26"]

    # Escalar metas generales a metas internacionales
    factor_int = TOTAL_META_INTERNACIONAL / meta_total_general
    metas_int = metas_general.copy()
    metas_int["metaSemanal"] = metas_int["metaSemanal"] * factor_int
    metas_int["metaDesgloce"] = metas_int["metaDesgloce"] * factor_int

    vg_int = base_weeks_int.merge(g2425_int, on="NoSemana", how="left") \
                           .merge(g2526_int, on="NoSemana", how="left") \
                           .merge(metas_int, on="NoSemana", how="left")

    vg_int["porcentajeMeta"] = (vg_int["venta25-26"] / vg_int["metaSemanal"] * 100).where(vg_int["metaSemanal"].ne(0))
    vg_int["porcentajeMetaDesgloce"] = (vg_int["25-26desgloce"] / vg_int["metaDesgloce"] * 100).where(vg_int["metaDesgloce"].ne(0))
    vg_int = vg_int[[
        "NoSemana", "venta24-25", "venta25-26", "metaSemanal",
        "24-25desgloce", "25-26desgloce", "metaDesgloce",
        "porcentajeMeta", "porcentajeMetaDesgloce"
    ]].sort_values("NoSemana")

    # ---------------- Embudos ----------------
    embudo_general = preparar_embudo_general(df_src)
    embudo_region = preparar_embudo_region(df_src)
    embudo_general_nacional = preparar_embudo_general_nacional(df_src)


    # ---------------- Escribir TODO ----------------
    with pd.ExcelWriter(work_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        ventas_2425.to_excel(writer, sheet_name="Ventas24-25", index=False)
        ventas_2526.to_excel(writer, sheet_name="Ventas25-26", index=False)
        region.to_excel(writer, sheet_name="VentasRegion", index=False)
        vg_all.to_excel(writer, sheet_name="VentasGeneral", index=False)
        vg_nat.to_excel(writer, sheet_name="VentasNacional", index=False)
        vg_int.to_excel(writer, sheet_name="VentasInternacional", index=False)
        embudo_general.to_excel(writer, sheet_name="EmbudoGeneral", index=False)
        embudo_region.to_excel(writer, sheet_name="EmbudoRegion", index=False)
        embudo_general_nacional.to_excel(writer, sheet_name="EmbudoGeneral_Nacional", index=False)


    wb = load_workbook(work_path)
    crear_o_reemplazar_tabla(wb, "Ventas24-25", "TablaVentas24-25")
    crear_o_reemplazar_tabla(wb, "Ventas25-26", "TablaVentas25-26")
    crear_o_reemplazar_tabla(wb, "VentasRegion", "TablaVentasRegion")
    crear_o_reemplazar_tabla(wb, "VentasGeneral", "TablaVentasGeneral")
    crear_o_reemplazar_tabla(wb, "VentasNacional", "TablaVentasNacional")
    crear_o_reemplazar_tabla(wb, "VentasInternacional", "TablaVentasInternacional")
    crear_o_reemplazar_tabla(wb, "EmbudoGeneral", "TablaEmbudoGeneral")
    crear_o_reemplazar_tabla(wb, "EmbudoRegion", "TablaEmbudoRegion")
    crear_o_reemplazar_tabla(wb, "EmbudoGeneral_Nacional", "TablaEmbudoGeneral_Nacional")

    wb.save(work_path)

    # Reemplazar archivo original
    copy_back_with_retries(work_path, dst)
    print("OK: Ventas24-25, Ventas25-26, VentasRegion, VentasGeneral, VentasNacional, VentasInternacional, EmbudoGeneral, EmbudoRegion, EmbudoGeneral_Nacional actualizadas.")

# ====== BLOQUE FINAL ======

def generar_modelodatos(
    base_dir: str,
    dst_path: str | None = None,
    src_path: str | None = None,
    vista_sheet: str = "Result 1",
):
    """
    Construye 'modeloDatos.xlsx' usando las rutas indicadas.
    Inyecta las rutas en las variables globales y ejecuta main().
    Devuelve la ruta final escrita.
    """
    global BASE_DIR, SRC_PATH, DST_PATH, SRC_SHEET

    # Inyectar rutas en las GLOBALS que usa el script
    BASE_DIR = base_dir
    SRC_PATH = src_path or os.path.join(BASE_DIR, "Tableau_WinRoom_ImagenFunnel_PBI.xlsx")
    DST_PATH = dst_path or os.path.join(BASE_DIR, "modeloDatos.xlsx")
    SRC_SHEET = vista_sheet

    print("⚙️ Construyendo modeloDatos.xlsx ...")
    main()  # reutiliza todo tu pipeline sin reescribirlo

    if not os.path.exists(DST_PATH):
        raise FileNotFoundError(
            f"No se encontró el archivo esperado:\n  {DST_PATH}\n"
            "Verifica que main() guardó con ese nombre/ruta."
        )

    print(f"✅ Generado en: {DST_PATH}")
    return DST_PATH


if __name__ == "__main__":
    # Solo GENERAR localmente (sin subir)
    generar_modelodatos()

