# -*- coding: utf-8 -*-
"""
h3h4.py
Genera/actualiza la hoja 'CategoriaAsesores' en modeloDatos.xlsx con la tabla
'TablaCategoriaAsesor' (Asesor, Categoria, Meta) a partir de:
- Seguimiento metas.xlsx (Nombre, Meta CRM (venta))  [primera hoja]
- modeloDatos.xlsx -> hoja 'Categorias' (Categoria, LimiteBajo, LimiteAlto)

Requisitos: pandas, openpyxl
"""

from __future__ import annotations
import os, re, time, math, shutil
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from typing import Optional

# Defaults de CONFIG (se sobreescriben con configure_paths)
# --- DEFAULTS (deben ir arriba, una sola vez) ---
BASE_DIR = os.getcwd()
MODELO_DATOS_PATH = os.path.join(BASE_DIR, "modeloDatos.xlsx")
WINROOM_PATH  = os.path.join(BASE_DIR, "Tableau_WinRoom_ImagenFunnel_PBI.xlsx")
WINROOM_SHEET = "Result 1"
SEGUIMIENTO_FILE = "Seguimiento metas.xlsx"

def configure_paths(base_dir: str):
    """Permite inyectar rutas desde import_export.main(origen)."""
    global BASE_DIR, MODELO_DATOS_PATH, WINROOM_PATH
    BASE_DIR = base_dir
    MODELO_DATOS_PATH = os.path.join(BASE_DIR, "modeloDatos.xlsx")
    WINROOM_PATH = os.path.join(BASE_DIR, "Tableau_WinRoom_ImagenFunnel_PBI.xlsx")

# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------
SEGUIMIENTO_FILE = "Seguimiento metas.xlsx"  # nombre EXACTO

HOJA_CATEGORIAS = "Categorias"
HOJA_SALIDA = "CategoriaAsesores"
TABLA_SALIDA = "TablaCategoriaAsesor"

COL_NOMBRE = "Nombre"
COL_META_CATEGORIA = "Meta ID"
COL_META_VENTA = "Meta CRM (venta)"
COL_REGION = "Región"
VERBOSE = False  # pon True si quieres ver avisos/diagnósticos
# Helpers
# ---------------------------------------------------------------------
def _to_float(x) -> float:
    if x is None:
        return np.nan
    if isinstance(x, (int, float, np.floating)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return np.nan
    s = re.sub(r"[^\d,.\-]", "", s)
    if s.count(",") > 0 and s.count(".") > 0:
        s = s.replace(",", "")
    elif s.count(",") > 0 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan

def _norm(s: object) -> str:
    # normaliza y reemplaza NBSP por espacio
    return "" if s is None else str(s).replace("\xa0", " ").strip().lower()

def _copiar_a_temporal(path_src: str) -> str:
    base, ext = os.path.splitext(path_src)
    tmp = f"{base}.__tmpread__{int(time.time())}{ext}"
    shutil.copyfile(path_src, tmp)
    return tmp

def _map_region(valor: object) -> str | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    s = str(valor).replace("\xa0", " ").strip()
    if s == "":
        return None
    pref = s[:3]  # <<-- SIN .upper()
    if pref == "CDM":
        return "MÉXICO"
    if pref == "Cen":
        return "CENTRO & OCCIDENTE"
    if pref == "MTY":
        return "MONTERREY"
    if pref == "Nor":
        return "NOROESTE"
    return None


# ---------------------------------------------------------------------
# Lecturas / transformaciones
# ---------------------------------------------------------------------
def _cargar_categorias(path_modelo: str, hoja: str = HOJA_CATEGORIAS) -> pd.DataFrame:
    if not os.path.exists(path_modelo):
        raise FileNotFoundError(f"No existe modeloDatos.xlsx: {path_modelo}")
    cat = pd.read_excel(path_modelo, sheet_name=hoja, engine="openpyxl")

    esperadas = {"Categoria", "LimiteBajo", "LimiteAlto"}
    faltan = esperadas - set(cat.columns)
    if faltan:
        raise ValueError(f"En la hoja '{hoja}' faltan columnas: {faltan}")

    def parse_alto(v):
        if pd.isna(v):
            return math.inf
        s = str(v).strip()
        if s in {"*", "", "inf", "Inf", "infinito", "Infinito"}:
            return math.inf
        return _to_float(s)

    cat["LimiteBajo"] = cat["LimiteBajo"].apply(_to_float)
    cat["LimiteAlto"] = cat["LimiteAlto"].apply(parse_alto)
    cat = cat.sort_values(["LimiteBajo", "LimiteAlto"]).reset_index(drop=True)
    return cat

def _validar_rangos(cat_df: pd.DataFrame) -> list[str]:
    problemas = []
    for i in range(len(cat_df) - 1):
        a, b = cat_df.iloc[i], cat_df.iloc[i + 1]
        if a["LimiteAlto"] < b["LimiteBajo"] - 1e-9:
            problemas.append(f"Hueco entre '{a.Categoria}' y '{b.Categoria}'")
        if a["LimiteAlto"] > b["LimiteBajo"] + 1e-9:
            problemas.append(f"Solape entre '{a.Categoria}' y '{b.Categoria}'")
    return problemas

def _clasificar_categoria(meta: float, cat_df: pd.DataFrame) -> str | None:
    if pd.isna(meta):
        return None
    for _, r in cat_df.iterrows():
        if meta >= r["LimiteBajo"] and meta <= r["LimiteAlto"]:
            return r["Categoria"]
    return None

def _leer_seguimiento(path_seguimiento: str) -> pd.DataFrame:
    """
    Lee 'Seguimiento metas.xlsx' aunque la fila 1 esté vacía y los encabezados
    estén desplazados. Devuelve columnas: Asesor, Meta, Region
    """
    if not os.path.exists(path_seguimiento):
        raise FileNotFoundError(f"No existe '{path_seguimiento}'")

    tmp_path = _copiar_a_temporal(path_seguimiento)
    try:
        raw = pd.read_excel(tmp_path, sheet_name=0, header=None, engine="openpyxl")

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

    raw = raw.replace("\xa0", " ", regex=True)

    # detectar fila de encabezados y posiciones de Nombre / Meta / Región
    """header_row_idx = None
    nombre_col_idx = None
    meta_col_idx = None
    region_col_idx = None"""

    header_row_idx = None
    nombre_col_idx = None
    meta_categoria_col_idx = None
    meta_venta_col_idx = None
    region_col_idx = None

    max_rows_to_scan = min(30, len(raw))

    for r in range(max_rows_to_scan):
        vals = [_norm(v) for v in raw.iloc[r].tolist()]
        def idx_of(label: str):
            try:
                return vals.index(_norm(label))
            except ValueError:
                return None

        """n_idx = idx_of(COL_NOMBRE)
        m_idx = idx_of(COL_META_VENTA)
        rg_idx = idx_of(COL_REGION)

        if (n_idx is not None) and (m_idx is not None) and (rg_idx is not None):
            header_row_idx = r
            nombre_col_idx = n_idx
            meta_col_idx = m_idx
            region_col_idx = rg_idx
            break"""
        
        n_idx = idx_of(COL_NOMBRE)
        mc_idx = idx_of(COL_META_CATEGORIA)
        mv_idx = idx_of(COL_META_VENTA)
        rg_idx = idx_of(COL_REGION)

        if (n_idx is not None) and (mc_idx is not None) and (mv_idx is not None) and (rg_idx is not None):
            header_row_idx = r
            nombre_col_idx = n_idx
            meta_categoria_col_idx = mc_idx
            meta_venta_col_idx = mv_idx
            region_col_idx = rg_idx
            break

    # fallback (según tu captura: fila 2 = índice 1; A = nombre, D = meta, B = región)
    """if header_row_idx is None:
        header_row_idx = 1
        nombre_col_idx = 0
        meta_col_idx = 3
        region_col_idx = 1"""
    if header_row_idx is None:
        header_row_idx = 1
        nombre_col_idx = 0
        region_col_idx = 1
        meta_categoria_col_idx = 2
        meta_venta_col_idx = 3

    #data = raw.iloc[header_row_idx + 1 :, [nombre_col_idx, meta_col_idx, region_col_idx]].copy()
    #data.columns = ["Asesor", "Meta", "Region"]
    data = raw.iloc[header_row_idx + 1 :,[nombre_col_idx, meta_categoria_col_idx, meta_venta_col_idx, region_col_idx]].copy()
    data.columns = ["Asesor", "MetaCategoria", "Meta", "Region"]

    # limpieza
    """data["Asesor"] = data["Asesor"].astype(str).str.strip()
    data["Meta"] = data["Meta"].apply(_to_float)
    data["Region"] = data["Region"].apply(_map_region)"""

    data["Asesor"] = data["Asesor"].astype(str).str.strip()
    data["MetaCategoria"] = data["MetaCategoria"].apply(_to_float)
    data["Meta"] = data["Meta"].apply(_to_float)
    data["Region"] = data["Region"].apply(_map_region)

    data = data.dropna(subset=["Asesor"])
    # si un asesor aparece repetido, nos quedamos con la meta máxima y la primera región no nula
    """data = (data.groupby("Asesor", as_index=False)
                 .agg(Meta=("Meta","max"),
                      Region=("Region","first")))"""
    data = (data.groupby("Asesor", as_index=False)
             .agg(MetaCategoria=("MetaCategoria", "max"),
                 Meta=("Meta", "max"),
                 Region=("Region", "first"),))
    return data


# ---------------------------------------------------------------------
# Escritura en modeloDatos.xlsx
# ---------------------------------------------------------------------
def _escribir_df_en_modelo(path_modelo: str, hoja: str, table_name: str, df: pd.DataFrame):
    if not os.path.exists(path_modelo):
        raise FileNotFoundError(f"No existe el archivo modeloDatos: {path_modelo}")

    if df.empty:
        if VERBOSE:
            print("No hay filas para escribir en 'CategoriaAsesores'. ¿Fuente vacía?")
        return

    wb = load_workbook(path_modelo)
    if hoja in wb.sheetnames:
        ws_old = wb[hoja]
        wb.remove(ws_old)
    ws = wb.create_sheet(hoja)

    # Escribir encabezados + datos
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Crear tabla
    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{get_column_letter(last_col)}{last_row}"

    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(path_modelo)



# ---------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------


def generar_categoria_asesores(
    base_dir: Optional[str] = None,
    modelo_path: Optional[str] = None,
    seguimiento_file: Optional[str] = None,
):
    """
    Construye 'CategoriaAsesores' -> 'TablaCategoriaAsesor' con columnas:
      Asesor, Region, Categoria, Meta (Meta en M MXN).
    - Lee 'seguimiento_file' (Asesor, Meta, Region mapeada) desde 'base_dir'
    - Clasifica la categoría según hoja 'Categorias' en 'modelo_path'
    """
    # 1) Resolver rutas y cargar rangos de Categorias
    if base_dir is None:
        base_dir = BASE_DIR  # fallback al configurado por configure_paths(...)
    if modelo_path is None:
        modelo_path = MODELO_DATOS_PATH
    if seguimiento_file is None:
        seguimiento_file = SEGUIMIENTO_FILE
    path_seguimiento = os.path.join(base_dir, seguimiento_file)
    cat_df = _cargar_categorias(modelo_path)

    # 2) Seguimiento (Asesor, Meta, Region)
    asesores = _leer_seguimiento(path_seguimiento)  # engine="openpyxl" ya dentro

    # 3) Clasificación por rangos de Categorias
    #asesores["Categoria"] = asesores["Meta"].apply(lambda m: _clasificar_categoria(m, cat_df))
    asesores["Categoria"] = asesores["MetaCategoria"].apply(lambda m: _clasificar_categoria(m, cat_df))

    # 4) Ordenar por categoría con el mismo orden que en la hoja 'Categorias'
    orden = cat_df["Categoria"].tolist()
    salida = asesores[["Asesor", "Region", "Categoria", "Meta"]].copy()
    salida["Categoria"] = pd.Categorical(salida["Categoria"], categories=orden, ordered=True)
    salida = salida.sort_values(["Categoria", "Asesor"], na_position="last").reset_index(drop=True)

    # 5) Escribir
    _escribir_df_en_modelo(modelo_path, HOJA_SALIDA, TABLA_SALIDA, salida)

    # 6) Aviso opcional
    fuera = salida["Categoria"].isna().sum()
    if VERBOSE and fuera:
        print(f"Aviso: {fuera} asesores sin categoría (revisa rangos en 'Categorias').")

    print(f"Listo: '{HOJA_SALIDA}' → '{TABLA_SALIDA}' en {modelo_path}")


# =======================
# UTIL: lectura WinRoom
# =======================
def _leer_winroom() -> pd.DataFrame:
    """
    Lee la exportación de WinRoom (Tableau_WinRoom_ImagenFunnel_PBI.xlsx, hoja WINROOM_SHEET)
    y normaliza:
      - 'USA' mapeado a 'INTERNACIONAL' en ZonaAgrupada
      - Filtra etapa 'Cerrada ganada'
      - Asegura tipos numéricos básicos
    """
    if not os.path.exists(WINROOM_PATH):
        raise FileNotFoundError(f"No encuentro WinRoom en: {WINROOM_PATH}")

    df = pd.read_excel(WINROOM_PATH, sheet_name=WINROOM_SHEET, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    # Normalizaciones mínimas
    if "ZonaAgrupada" in df.columns:
        df["ZonaAgrupada"] = df["ZonaAgrupada"].astype(str).str.strip()
        df.loc[df["ZonaAgrupada"].str.upper().eq("USA"), "ZonaAgrupada"] = "INTERNACIONAL"

    # Filtros y tipos esperados
    if "Etapa" in df.columns:
        df = df[df["Etapa"].astype(str).str.strip().str.lower() == "cerrada ganada"]

    # Tipos
    if "NoSemana" in df.columns:
        df["NoSemana"] = pd.to_numeric(df["NoSemana"], errors="coerce").fillna(0).astype(int)
    if "Importe" in df.columns:
        df["Importe"] = pd.to_numeric(df["Importe"], errors="coerce").fillna(0.0)

    # Asegurar columnas clave existan
    requeridas = {"NoSemana", "Periodo", "ZonaAgrupada", "Asesor", "Importe"}
    faltan = requeridas - set(df.columns)
    if faltan:
        raise ValueError(f"En WinRoom faltan columnas: {sorted(faltan)}")

    return df[["NoSemana", "Periodo", "ZonaAgrupada", "Asesor", "Etapa", "Importe"]]


# ===============================================
# BUILDER: VentasAsesoresRegion (según especificación)
# ===============================================
def generar_ventas_asesores_region(
    modelo_path: str = MODELO_DATOS_PATH,
    hoja_salida: str = "VentasAsesoresRegion",
    tabla_salida: str = "TablaVentasAsesorRegion",
):
    # 1) WinRoom
    wr = _leer_winroom()  # columnas: NoSemana, Periodo, ZonaAgrupada (USA->INTERNACIONAL), Asesor, Etapa, Importe

    # 2) Ventas acumuladas por Asesor–Zona–Semana–Periodo
    agg = (wr.groupby(["Asesor", "ZonaAgrupada", "NoSemana", "Periodo"], as_index=False)
             .agg(Importe=("Importe", "sum")))

    p2425 = agg[agg["Periodo"] == "24-25"][["Asesor","ZonaAgrupada","NoSemana","Importe"]].rename(columns={"Importe":"Venta24_25_acum"})
    p2526 = agg[agg["Periodo"] == "25-26"][["Asesor","ZonaAgrupada","NoSemana","Importe"]].rename(columns={"Importe":"Venta25_26_acum"})

    # Presencia en WinRoom (Aparece=1)
    pres = (agg.groupby(["Asesor","ZonaAgrupada","NoSemana"], as_index=False)
               .size().rename(columns={"size":"Aparece"}))
    pres["Aparece"] = 1

    ventas = p2425.merge(p2526, on=["Asesor","ZonaAgrupada","NoSemana"], how="outer").merge(
        pres, on=["Asesor","ZonaAgrupada","NoSemana"], how="left"
    )

    # 3) Rejilla asesores × semanas
    semanas = np.sort(wr["NoSemana"].dropna().astype(int).unique())
    df_semanas = pd.DataFrame({"NoSemana": semanas})

    # CategoriaAsesores (con Region si existe)
    cats = pd.read_excel(modelo_path, sheet_name="CategoriaAsesores", engine="openpyxl")
    cats.columns = [str(c).strip() for c in cats.columns]
    if "Region" not in cats.columns:
        cats["Region"] = None
    cats = cats[["Asesor","Region","Categoria","Meta"]].copy()
    cats["Meta"] = cats["Meta"].apply(_to_float).fillna(0.0)  # M MXN

    # Zona por asesor: WinRoom si existe; si nunca, Region de CategoriaAsesores
    pares_wz = wr[["Asesor","ZonaAgrupada"]].drop_duplicates()
    asesores = (cats[["Asesor","Region"]].drop_duplicates().merge(pares_wz, on="Asesor", how="left"))
    asesores["ZonaAgrupada"] = asesores["ZonaAgrupada"].fillna(asesores["Region"])
    asesores["ZonaAgrupada"] = asesores["ZonaAgrupada"].fillna("SIN ZONA")

    asesores["key"] = 1; df_semanas["key"] = 1
    rejilla = asesores.merge(df_semanas, on="key").drop(columns="key")

    # 4) Unir ventas a la rejilla
    base = rejilla.merge(ventas, on=["Asesor","ZonaAgrupada","NoSemana"], how="left")
    base[["Venta24_25_acum","Venta25_26_acum"]] = base[["Venta24_25_acum","Venta25_26_acum"]].fillna(0.0)
    base["Aparece"] = base["Aparece"].fillna(0).astype(int)

    base["NoSemana"] = pd.to_numeric(base["NoSemana"], errors="coerce").fillna(0).astype(int)
    base = base.sort_values(["Asesor","ZonaAgrupada","NoSemana"]).reset_index(drop=True)

    # 5) Columnas finales de ventas (acumulado real WinRoom)
    base["Venta24-25"] = base["Venta24_25_acum"]
    base["Venta25-26"] = base["Venta25_26_acum"]

    # Desgloses = diferencia vs semana anterior por Asesor–Zona
    base["Venta24-25desgloce"] = base.groupby(["Asesor","ZonaAgrupada"])["Venta24-25"].diff().fillna(base["Venta24-25"])
    base["Venta25-26desgloce"] = base.groupby(["Asesor","ZonaAgrupada"])["Venta25-26"].diff().fillna(base["Venta25-26"])

    # 6) Categoría y Meta
    base = base.merge(cats[["Asesor","Categoria","Meta"]], on="Asesor", how="left", validate="m:1")
    miss = base["Categoria"].isna()
    base.loc[miss & (base["ZonaAgrupada"].astype(str).str.upper()=="INTERNACIONAL"), "Categoria"] = "INTERNACIONAL"
    base.loc[miss & (base["ZonaAgrupada"].astype(str).str.upper()!="INTERNACIONAL"), "Categoria"] = "INCOMPATIBLE"

    # 7) Metas: semanal y acumulada por semana N
    base["Meta_MXN"] = base["Meta"] * 1_000_000.0
    base["MetaDesgloce"] = np.where(base["Meta_MXN"]>0, base["Meta_MXN"]/52.0, 0.0)
    base["Meta"] = (base["MetaDesgloce"] * base["NoSemana"]).astype(float)

    # 8) Porcentajes
    base["PorcentajeMeta"] = np.where(base["Meta"]>0, (base["Venta25-26"]/base["Meta"])*100.0, 0.0)
    base["ProcentajeDesgloce"] = np.where(base["MetaDesgloce"]>0, (base["Venta25-26desgloce"]/base["MetaDesgloce"])*100.0, 0.0)

    # 9) Salida
    salida = base[[
        "NoSemana","ZonaAgrupada","Asesor","Categoria",
        "Venta24-25","Venta25-26","Meta","PorcentajeMeta",
        "Venta24-25desgloce","Venta25-26desgloce","MetaDesgloce","ProcentajeDesgloce",
        "Aparece",
    ]].sort_values(["ZonaAgrupada","Asesor","NoSemana"]).reset_index(drop=True)

    _escribir_df_en_modelo(modelo_path, hoja_salida, tabla_salida, salida)



def generar_ventas_asesores_general(
    modelo_path: str = MODELO_DATOS_PATH,
    hoja_salida: str = "VentasAsesoresGeneral",
    tabla_salida: str = "TablaVentasAsesorGeneral",
):
    wr = _leer_winroom()

    # Categoría por asesor
    cats = pd.read_excel(modelo_path, sheet_name="CategoriaAsesores", engine="openpyxl")
    cats.columns = [str(c).strip() for c in cats.columns]
    cats = cats[["Asesor","Categoria","Meta"]].copy()
    cats["Meta"] = cats["Meta"].apply(_to_float).fillna(0.0)

    wr_cat = wr.merge(cats[["Asesor","Categoria"]], on="Asesor", how="left", validate="m:1")
    miss = wr_cat["Categoria"].isna()
    wr_cat.loc[miss & (wr_cat["ZonaAgrupada"].str.upper()=="INTERNACIONAL"), "Categoria"] = "INTERNACIONAL"
    wr_cat.loc[miss & (wr_cat["ZonaAgrupada"].str.upper()!="INTERNACIONAL"), "Categoria"] = "INCOMPATIBLE"

    # Ventas acumuladas de WinRoom por semana y categoría (sumar asesores)
    agg = (wr_cat.groupby(["NoSemana","Categoria","Periodo"], as_index=False)
                 .agg(Importe=("Importe","sum")))

    v2425 = agg[agg["Periodo"]=="24-25"][["NoSemana","Categoria","Importe"]].rename(columns={"Importe":"Venta24_25_acum"})
    v2526 = agg[agg["Periodo"]=="25-26"][["NoSemana","Categoria","Importe"]].rename(columns={"Importe":"Venta25_26_acum"})
    base = pd.merge(v2425, v2526, on=["NoSemana","Categoria"], how="outer").fillna(0.0)

    base["NoSemana"] = pd.to_numeric(base["NoSemana"], errors="coerce").fillna(0).astype(int)
    base = base.sort_values(["Categoria","NoSemana"]).reset_index(drop=True)

    # Acumulado ya viene listo → solo renombrar a columnas finales
    base["Venta24-25"] = base["Venta24_25_acum"]
    base["Venta25-26"] = base["Venta25_26_acum"]

    # Desgloses por categoría
    base["Venta24-25desgloce"] = base.groupby(["Categoria"])["Venta24-25"].diff().fillna(base["Venta24-25"])
    base["Venta25-26desgloce"] = base.groupby(["Categoria"])["Venta25-26"].diff().fillna(base["Venta25-26"])

    # Meta por categoría SOLO desde CategoriaAsesores (evita duplicados)
    meta_cat_mm = cats.groupby("Categoria", as_index=False)["Meta"].sum()
    meta_cat_mm.loc[meta_cat_mm["Categoria"].isin(["INTERNACIONAL","INCOMPATIBLE"]), "Meta"] = 0.0

    base = base.merge(meta_cat_mm, on="Categoria", how="left")
    base["Meta"] = base["Meta"].fillna(0.0)
    base["Meta_MXN"] = base["Meta"] * 1_000_000.0

    base["PorcentajeMeta"] = np.where(base["Meta_MXN"]>0, (base["Venta25-26"]/base["Meta_MXN"])*100.0, 0.0)
    base["MetaDesgloce"]   = np.where(base["Meta_MXN"]>0, base["Meta_MXN"]/52.0, 0.0)
    base["ProcentajeDesgloce"] = np.where(base["MetaDesgloce"]>0, (base["Venta25-26desgloce"]/base["MetaDesgloce"])*100.0, 0.0)

    base["ZonaAgrupada"] = "GENERAL"
    base["Asesor"] = "GENERAL"

    salida = base[[
        "NoSemana","ZonaAgrupada","Asesor","Categoria",
        "Venta24-25","Venta25-26","Meta","PorcentajeMeta",
        "Venta24-25desgloce","Venta25-26desgloce","MetaDesgloce","ProcentajeDesgloce",
    ]].sort_values(["Categoria","NoSemana"]).reset_index(drop=True)

    _escribir_df_en_modelo(modelo_path, hoja_salida, tabla_salida, salida)


def generar_ventas_asesores_nacional(
    modelo_path: str = MODELO_DATOS_PATH,
    hoja_salida: str = "VentasAsesoresNacional",
    tabla_salida: str = "TablaVentasAsesorNacional",
):
    wr = _leer_winroom()
    wr_nac = wr[wr["ZonaAgrupada"].str.upper() != "INTERNACIONAL"].copy()

    cats = pd.read_excel(modelo_path, sheet_name="CategoriaAsesores", engine="openpyxl")
    cats.columns = [str(c).strip() for c in cats.columns]
    cats = cats[["Asesor","Categoria","Meta"]].copy()
    cats["Meta"] = cats["Meta"].apply(_to_float).fillna(0.0)

    wr_cat = wr_nac.merge(cats[["Asesor","Categoria"]], on="Asesor", how="left", validate="m:1")
    wr_cat.loc[wr_cat["Categoria"].isna(), "Categoria"] = "INCOMPATIBLE"

    # Ventas acumuladas por semana/categoría/periodo (solo nacional)
    agg = (wr_cat.groupby(["NoSemana","Categoria","Periodo"], as_index=False)
                 .agg(Importe=("Importe","sum")))
    v2425 = agg[agg["Periodo"]=="24-25"][["NoSemana","Categoria","Importe"]].rename(columns={"Importe":"Venta24_25_acum"})
    v2526 = agg[agg["Periodo"]=="25-26"][["NoSemana","Categoria","Importe"]].rename(columns={"Importe":"Venta25_26_acum"})
    base = pd.merge(v2425, v2526, on=["NoSemana","Categoria"], how="outer").fillna(0.0)

    base["NoSemana"] = pd.to_numeric(base["NoSemana"], errors="coerce").fillna(0).astype(int)
    base = base.sort_values(["Categoria","NoSemana"]).reset_index(drop=True)

    base["Venta24-25"] = base["Venta24_25_acum"]
    base["Venta25-26"] = base["Venta25_26_acum"]

    base["Venta24-25desgloce"] = base.groupby(["Categoria"])["Venta24-25"].diff().fillna(base["Venta24-25"])
    base["Venta25-26desgloce"] = base.groupby(["Categoria"])["Venta25-26"].diff().fillna(base["Venta25-26"])

    # Meta por categoría (solo asesores que existen en wr_nac – nacional)
    asesores_nac = wr_nac[["Asesor"]].drop_duplicates()
    cats_nac = asesores_nac.merge(cats, on="Asesor", how="left")
    meta_cat_mm = cats_nac.groupby("Categoria", as_index=False)["Meta"].sum()
    meta_cat_mm.loc[meta_cat_mm["Categoria"]=="INCOMPATIBLE", "Meta"] = 0.0

    base = base.merge(meta_cat_mm, on="Categoria", how="left")
    base["Meta"] = base["Meta"].fillna(0.0)
    base["Meta_MXN"] = base["Meta"] * 1_000_000.0

    base["PorcentajeMeta"] = np.where(base["Meta_MXN"]>0, (base["Venta25-26"]/base["Meta_MXN"])*100.0, 0.0)
    base["MetaDesgloce"]   = np.where(base["Meta_MXN"]>0, base["Meta_MXN"]/52.0, 0.0)
    base["ProcentajeDesgloce"] = np.where(base["MetaDesgloce"]>0, (base["Venta25-26desgloce"]/base["MetaDesgloce"])*100.0, 0.0)

    base["ZonaAgrupada"] = "NACIONAL"
    base["Asesor"] = "NACIONAL"

    salida = base[[
        "NoSemana","ZonaAgrupada","Asesor","Categoria",
        "Venta24-25","Venta25-26","Meta","PorcentajeMeta",
        "Venta24-25desgloce","Venta25-26desgloce","MetaDesgloce","ProcentajeDesgloce",
    ]].sort_values(["Categoria","NoSemana"]).reset_index(drop=True)

    _escribir_df_en_modelo(modelo_path, hoja_salida, tabla_salida, salida)




def generar_actualizacion(
    modelo_path: str = MODELO_DATOS_PATH,
    hoja_salida: str = "Actualizacion",
    tabla_salida: str = "TablaActualizacion",
):
    """
    Crea/actualiza la hoja 'Actualizacion' con la tabla 'TablaActualizacion'
    (columna única 'Fecha') usando la fecha y hora actuales.
    """
    # Usa objeto datetime (Excel lo reconoce como fecha/hora)
    now_dt = datetime.now()
    df = pd.DataFrame([{"Fecha": now_dt}])

    _escribir_df_en_modelo(modelo_path, hoja_salida, tabla_salida, df)

# ---------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------
if __name__ == "__main__":
    import argparse, os
    try:
        # usamos rutas.json / rutas.py si está disponible
        from rutas import get_paths
    except ImportError:
        get_paths = None

    parser = argparse.ArgumentParser(description="Genera tablas HU3/HU4 en modeloDatos.xlsx")
    parser.add_argument(
        "--origen",
        choices=["Andrea", "Antonio", "Erik"],
        default=os.getenv("U4O_ORIGEN"),
        help="Origen de rutas; si se omite, se usa el DEFAULT de rutas.py (si existe).",
    )
    parser.add_argument(
        "--todo",
        action="store_true",
        help="Si se indica, genera todas las tablas (categorías, ventas region/general/nacional, cumplimiento, actualización).",
    )
    args = parser.parse_args()

    # Resolver BASE_DIR
    if get_paths is not None:
        BASE_DIR, _ = get_paths(args.origen)
    else:
        # fallback: carpeta actual
        BASE_DIR = os.getcwd()

    # Apuntar rutas del módulo
    configure_paths(BASE_DIR)

    if args.todo:
        generar_categoria_asesores()
        generar_ventas_asesores_region()
        generar_ventas_asesores_general()
        generar_ventas_asesores_nacional()
        generar_actualizacion()
        print("✅ HU3/HU4 listas.")
    else:
        # ejemplo mínimo: solo categorías
        generar_categoria_asesores()
        print("✅ TablaCategoriaAsesor generada.")
