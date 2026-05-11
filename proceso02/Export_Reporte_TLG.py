# Export_Reporte_TLG.py
import os
import warnings
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from Conexion_SQL import crear_conexion_sql


_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_OUTPUT_DIR = os.path.join(_BASE_DIR, "output")

OUTPUT_DIR_BY_ORIGEN = {
    "Antonio": _OUTPUT_DIR,
    "Erik":    _OUTPUT_DIR,
}
META_GENERAL = 50_000_000  # $50,000,000 meta general
META_ZONA = 10_000_000  # $10,000,000 por ZonaAgrupada

PROGRAMAS_COLUMNS = [
    "FechaCierre",
    "AñoMeta",
    "Escuela",
    "ZonaAgrupada",
    "CampusUsuario",
    "Asesor",
    "TipoPrograma",
    "TipoRegistro",
    "Empresa",
    "NombreOportunidad",
    "Importe",
    "Duración",
    "Participantes",
    "Coordinador",
    "Diseñador",
    "IdOportunidad",
    "AreaTematica",
    "TipoIniciativa",
    "InicioEjecucion",
    "TerminoEjecucion",
    "NivelImpacto",
    "ComponenteTLG",
    "PorcentajeTLG",
    "ImporteTLG",
    "ImporteNoTLG",
]

DETALLES_COLUMNS = [
    "IdOportunidad",
    "TLG_Nomenclatura",
    "TLG_Tipo_Componente",
    "TLG_Area",
    "TLG_Trayectoria",
    "TLG_Competencia",
    "TLG_Subcompetencia",
]

def get_output_dir(origen: str) -> str:
    origen_norm = (origen or "").strip()
    if origen_norm in OUTPUT_DIR_BY_ORIGEN:
        return OUTPUT_DIR_BY_ORIGEN[origen_norm]
    raise ValueError(f"Origen no válido: '{origen}'. Usa 'Antonio' o 'Erik'.")

def obtener_datos(anio_meta: str = "2025-2026") -> pd.DataFrame:
    query = f"""
    SELECT
        sth.[FechaCierre],
        sth.[AñoMeta],
        sth.[Escuela],
        sth.[ZonaAgrupada],
        sth.[CampusUsuario],
        sth.[PropietarioOportunidad] AS [Asesor],
        sth.[TipoPrograma],
        sth.[TipoRegistro],
        sth.[EmpresaCRM] AS [Empresa],
        sth.[NombreOportunidad],
        sth.[Importe],
        sth.[Duración],
        sth.[Participantes],
        sth.[Coordinador],
        sth.[Diseñador],
        sth.[IdOportunidad],
        sth.[AreaTematica],
        sth.[TipoIniciativa],
        sth.[InicioEjecucion],
        sth.[TerminoEjecucion],
        sth.[NivelImpacto],
        sth.[ComponenteTLG],
        sth.[PorcentajeTLG],
        sth.[ImporteTLG],
        sth.[ImporteNoTLG],

        lc.[TLG_Nomenclatura],
        lc.[TLG_Tipo_Componente],
        lc.[TLG_Area],
        lc.[TLG_Trayectoria],
        lc.[TLG_Competencia],
        lc.[TLG_Subcompetencia]

    FROM Edata_Qlik.dbo.Tableau_U4O_Historico AS sth
    LEFT JOIN Edata_Qlik.dbo.ETL_Extraer_TLG_Componente AS ec
        ON sth.[IdOportunidad] = ec.[Id]
    LEFT JOIN Edata_Qlik.dbo.ETL_TLG_ListaComponentes AS lc
        ON ec.[Nomenclatura] = lc.[TLG_Nomenclatura]
    WHERE
        sth.[EtapaDetalle] = 'Cerrada ganada'
        AND sth.[AñoMeta] = '{anio_meta}'
        AND sth.[ComponenteTLG] = 'true';
    """

    conn = crear_conexion_sql('Edata_Qlik')
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="pandas only supports SQLAlchemy connectable*",
                category=UserWarning
            )
            df = pd.read_sql_query(query, conn)
        return df
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _asegurar_columnas(df: pd.DataFrame, columnas: list[str]) -> pd.DataFrame:
    for col in columnas:
        if col not in df.columns:
            df[col] = None
    return df[columnas].copy()


def aplicar_formato_tec(
    ws,
    nombre_tabla: str,
    currency_cols: list[str] | None = None,
    date_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
    tec_blue: str = "0039A6",
    stripe: str = "EAF2FF"
):
    currency_cols = currency_cols or []
    date_cols = date_cols or []
    percent_cols = percent_cols or []

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=nombre_tabla, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    header_fill = PatternFill("solid", fgColor=tec_blue)
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    stripe_fill = PatternFill("solid", fgColor=stripe)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for r in range(2, max_row + 1):
        fill = stripe_fill if (r % 2 == 0) else white_fill
        for c in range(1, max_col + 1):
            ws.cell(r, c).fill = fill

    headers = {ws.cell(1, c).value: c for c in range(1, max_col + 1)}

    mxn_format = '[$$-es-MX]#,##0.00'
    date_format = "dd/mm/yyyy"
    pct_format = "0.00%"

    for col_name in currency_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for r in range(2, max_row + 1):
                cell = ws.cell(r, col_idx)
                cell.number_format = mxn_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_name in date_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for r in range(2, max_row + 1):
                cell = ws.cell(r, col_idx)
                cell.number_format = date_format
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_name in percent_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for r in range(2, max_row + 1):
                cell = ws.cell(r, col_idx)
                cell.number_format = pct_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in range(1, max_col + 1):
        max_len = 0
        for r in range(1, max_row + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            if hasattr(val, "strftime"):
                length = 10
            else:
                length = len(str(val))
            max_len = max(max_len, length)

        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(c)].width = width


def aplicar_formato_tec_rango(
    ws,
    nombre_tabla: str,
    start_row: int,
    start_col: int,
    nrows: int,
    ncols: int,
    currency_cols: list[str] | None = None,
    date_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
    tec_blue: str = "0039A6",
    stripe: str = "EAF2FF"
):
    currency_cols = currency_cols or []
    date_cols = date_cols or []
    percent_cols = percent_cols or []

    if nrows < 1 or ncols < 1:
        return

    end_row = start_row + nrows - 1
    end_col = start_col + ncols - 1

    start_cell = f"{get_column_letter(start_col)}{start_row}"
    end_cell = f"{get_column_letter(end_col)}{end_row}"
    ref = f"{start_cell}:{end_cell}"

    # Si solo hay encabezado (sin filas), no creamos tabla para evitar errores
    if nrows >= 2:
        table = Table(displayName=nombre_tabla, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    header_fill = PatternFill("solid", fgColor=tec_blue)
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[start_row].height = 22
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    stripe_fill = PatternFill("solid", fgColor=stripe)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for r in range(start_row + 1, end_row + 1):
        fill = stripe_fill if (r % 2 == 0) else white_fill
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).fill = fill

    headers = {ws.cell(start_row, c).value: c for c in range(start_col, end_col + 1)}

    mxn_format = '[$$-es-MX]#,##0.00'
    date_format = "dd/mm/yyyy"
    pct_format = "0.00%"

    for col_name in currency_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for r in range(start_row + 1, end_row + 1):
                cell = ws.cell(r, col_idx)
                cell.number_format = mxn_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_name in date_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for r in range(start_row + 1, end_row + 1):
                cell = ws.cell(r, col_idx)
                cell.number_format = date_format
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_name in percent_cols:
        if col_name in headers:
            col_idx = headers[col_name]
            for r in range(start_row + 1, end_row + 1):
                cell = ws.cell(r, col_idx)
                cell.number_format = pct_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in range(start_col, end_col + 1):
        max_len = 0
        for r in range(start_row, end_row + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            if hasattr(val, "strftime"):
                length = 10
            else:
                length = len(str(val))
            max_len = max(max_len, length)

        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(c)].width = width


def _titulo(ws, cell: str, texto: str, tec_blue: str = "0039A6", size: int = 13):
    c = ws[cell]
    c.value = texto
    c.font = Font(bold=True, color=tec_blue, size=size)
    c.alignment = Alignment(horizontal="left", vertical="center")


def exportar_excel(df: pd.DataFrame, output_dir: str = OUTPUT_DIR_BY_ORIGEN) -> str:
    os.makedirs(output_dir, exist_ok=True)

    filename = "Reporte_TLG.xlsx"
    filepath = os.path.join(output_dir, filename)

    # --- Programas ---
    df_programas = _asegurar_columnas(df, PROGRAMAS_COLUMNS)

    # --- Limpieza: USA pertenece a INTERNACIONAL ---
    df_programas["ZonaAgrupada"] = df_programas["ZonaAgrupada"].fillna("").astype(str).str.strip()
    df_programas.loc[df_programas["ZonaAgrupada"].str.upper() == "USA", "ZonaAgrupada"] = "INTERNACIONAL"
    df_programas.loc[df_programas["ZonaAgrupada"] == "", "ZonaAgrupada"] = None

    for col in ["FechaCierre", "InicioEjecucion", "TerminoEjecucion"]:
        if col in df_programas.columns:
            df_programas[col] = pd.to_datetime(df_programas[col], errors="coerce")

    # Programas sin repetir IdOportunidad (más reciente por FechaCierre)
    if "IdOportunidad" in df_programas.columns:
        if "FechaCierre" in df_programas.columns:
            df_programas = df_programas.sort_values(
                by=["IdOportunidad", "FechaCierre"],
                ascending=[True, False],
                na_position="last"
            )
        else:
            df_programas = df_programas.sort_values(by=["IdOportunidad"])
        df_programas = df_programas.drop_duplicates(subset=["IdOportunidad"], keep="first")

    # --- Detalles ---
    df_detalles = _asegurar_columnas(df, DETALLES_COLUMNS)

    # --- Actualizacion ---
    df_actualizacion = pd.DataFrame({
        "Fecha de actualización": [datetime.now()]
    })

    # --- Resumen (tablas) ---
    df_prog_sum = df_programas.copy()
    df_prog_sum["ZonaAgrupada"] = df_prog_sum["ZonaAgrupada"].fillna("Sin zona")
    df_prog_sum["Asesor"] = df_prog_sum["Asesor"].fillna("Sin asesor")
    df_prog_sum["ImporteTLG"] = pd.to_numeric(df_prog_sum["ImporteTLG"], errors="coerce").fillna(0)

    # 1) ZonaAgrupada
    df_res_zona = (
        df_prog_sum.groupby("ZonaAgrupada", as_index=False)["ImporteTLG"]
        .sum()
        .rename(columns={"ImporteTLG": "Acumulado ImporteTLG"})
    )
    df_res_zona["Meta"] = META_ZONA
    df_res_zona["Avance %"] = df_res_zona["Acumulado ImporteTLG"] / df_res_zona["Meta"]
    df_res_zona = df_res_zona.sort_values("Acumulado ImporteTLG", ascending=False)

    # 2) Asesor
    df_res_asesor = (
        df_prog_sum.groupby("Asesor", as_index=False)
        .agg(
            Programas=("IdOportunidad", "nunique"),
            **{"Acumulado ImporteTLG": ("ImporteTLG", "sum")}
        )
        .sort_values("Acumulado ImporteTLG", ascending=False)
    )

    # 3) Distribución Detalles
    df_det_sum = df_detalles.copy()
    for c in ["TLG_Tipo_Componente", "TLG_Area", "TLG_Trayectoria"]:
        df_det_sum[c] = df_det_sum[c].fillna("Sin dato")

    df_res_dist = (
        df_det_sum.groupby(["TLG_Tipo_Componente", "TLG_Area", "TLG_Trayectoria"], as_index=False)
        .size()
        .rename(columns={"size": "Registros"})
        .sort_values("Registros", ascending=False)
    )
    total_reg = float(df_res_dist["Registros"].sum()) if len(df_res_dist) else 0.0
    df_res_dist["%Distribución"] = (df_res_dist["Registros"] / total_reg) if total_reg else 0.0

    # 4) Avance general
    total_importe_tlg = float(df_prog_sum["ImporteTLG"].sum())

    df_res_general = pd.DataFrame({
        "Acumulado ImporteTLG": [total_importe_tlg],
        "Meta General": [META_GENERAL],
        "Avance %": [total_importe_tlg / META_GENERAL if META_GENERAL else 0],
    })

    # --- Exporta ---
    try:
        with pd.ExcelWriter(filepath, engine="openpyxl", mode="w") as writer:
            # 1) Programas
            df_programas.to_excel(writer, index=False, sheet_name="Programas")

            # 2) Detalles
            df_detalles.to_excel(writer, index=False, sheet_name="Detalles")

            # 3) Resumen (con tablas acomodadas)
            sheet_res = "Resumen"

            # --- Tabla 0: Resumen General (arriba) ---
            general_title_row = 2
            general_table_row = 3
            df_res_general.to_excel(writer, index=False, sheet_name=sheet_res, startrow=general_table_row - 1, startcol=0)

            last_row_general = general_table_row + len(df_res_general)  # header + data

            # --- Tabla 1: Zona ---
            zona_title_row = last_row_general + 2
            zona_table_row = zona_title_row + 1
            df_res_zona.to_excel(writer, index=False, sheet_name=sheet_res, startrow=zona_table_row - 1, startcol=0)

            last_row_zona = zona_table_row + len(df_res_zona)

            # --- Tabla 2: Asesor ---
            asesor_title_row = last_row_zona + 2
            asesor_table_row = asesor_title_row + 1
            df_res_asesor.to_excel(writer, index=False, sheet_name=sheet_res, startrow=asesor_table_row - 1, startcol=0)

            last_row_asesor = asesor_table_row + len(df_res_asesor)

            # --- Tabla 3: Distribución ---
            dist_title_row = last_row_asesor + 2
            dist_table_row = dist_title_row + 1
            df_res_dist.to_excel(writer, index=False, sheet_name=sheet_res, startrow=dist_table_row - 1, startcol=0)

            # 4) Actualizacion
            df_actualizacion.to_excel(writer, index=False, sheet_name="Actualizacion")

            # Hojas (ya quedan en orden: Programas, Detalles, Resumen, Actualizacion)
            ws_prog = writer.sheets["Programas"]
            ws_det  = writer.sheets["Detalles"]
            ws_res  = writer.sheets["Resumen"]
            ws_act  = writer.sheets["Actualizacion"]

            # Formato hojas base
            aplicar_formato_tec(
                ws_prog,
                nombre_tabla="tblProgramas",
                currency_cols=["Importe", "ImporteTLG", "ImporteNoTLG"],
                date_cols=["FechaCierre", "InicioEjecucion", "TerminoEjecucion"],
            )

            aplicar_formato_tec(
                ws_det,
                nombre_tabla="tblDetalles",
            )

            aplicar_formato_tec(
                ws_act,
                nombre_tabla="tblActualizacion",
                date_cols=["Fecha de actualización"],
            )

            # Resumen: títulos
            _titulo(ws_res, "A1", "Resumen TLG", size=16)
            _titulo(ws_res, f"A{general_title_row}", "Resumen general (Meta: $50,000,000)", size=13)
            _titulo(ws_res, f"A{zona_title_row}", "Acumulado por ZonaAgrupada (Meta: $10,000,000)", size=13)
            _titulo(ws_res, f"A{asesor_title_row}", "Acumulado por Asesor", size=13)
            _titulo(ws_res, f"A{dist_title_row}", "Distribución de Componentes (Tipo x Área x Trayectoria)", size=13)

            # Resumen: formato por rango (tablas)
            aplicar_formato_tec_rango(
            ws_res,
            nombre_tabla="tblResumenGeneral",
            start_row=general_table_row,
            start_col=1,
            nrows=len(df_res_general) + 1,
            ncols=df_res_general.shape[1],
            currency_cols=["Acumulado ImporteTLG", "Meta General"],
            percent_cols=["Avance %"],
            )

            aplicar_formato_tec_rango(
                ws_res,
                nombre_tabla="tblResumenZona",
                start_row=zona_table_row,
                start_col=1,
                nrows=len(df_res_zona) + 1,
                ncols=df_res_zona.shape[1],
                currency_cols=["Acumulado ImporteTLG", "Meta"],
                percent_cols=["Avance %"],
            )

            aplicar_formato_tec_rango(
                ws_res,
                nombre_tabla="tblResumenAsesor",
                start_row=asesor_table_row,
                start_col=1,
                nrows=len(df_res_asesor) + 1,
                ncols=df_res_asesor.shape[1],
                currency_cols=["Acumulado ImporteTLG"],
            )

            aplicar_formato_tec_rango(
                ws_res,
                nombre_tabla="tblResumenDistrib",
                start_row=dist_table_row,
                start_col=1,
                nrows=len(df_res_dist) + 1,
                ncols=df_res_dist.shape[1],
                percent_cols=["%Distribución"],
            )

            # Freeze panes
            for sheet_name in ["Programas", "Detalles", "Actualizacion"]:
                writer.sheets[sheet_name].freeze_panes = "A2"
            ws_res.freeze_panes = "A5"

    except PermissionError:
        raise RuntimeError(
            f"El archivo está abierto y no se puede sobrescribir:\n{filepath}\n"
            "Ciérralo en Excel y vuelve a correr el proceso."
        )

    return filepath


def main(origen: str = "Antonio", anio_meta: str = "2025-2026", output_dir: str = OUTPUT_DIR_BY_ORIGEN):

    output_dir = get_output_dir(origen)
    
    print(f"[Export_Reporte_TLG] Iniciando. Origen={origen}, AñoMeta={anio_meta}")
    print(f"[Export_Reporte_TLG] OutputDir={output_dir}")

    df = obtener_datos(anio_meta=anio_meta)
    print(f"[Export_Reporte_TLG] Registros obtenidos: {len(df):,}")

    path = exportar_excel(df, output_dir=output_dir)
    print(f"[Export_Reporte_TLG] Archivo generado: {path}")


if __name__ == "__main__":
    main()